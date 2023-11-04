import io
import datetime

import calendar
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side

from Exceptions import NoWorkers, WorkerNotHaveWorkHours
from logger import Logger
from tables.Models import Worker, Employer

cls_logger = Logger()
logger = cls_logger.get_logger


class AppearanceOTWHSheet:
    """SHEETS FOR THE APPEARANCE OF THE WORKING HOUR"""

    def __init__(self, employer: Employer):
        self.Employer: Employer = employer
        self.creation_date: datetime = datetime.datetime.now()
        self.start_writing_period: datetime = datetime.date(self.creation_date.year, self.creation_date.month, 1)
        self.start_billing_period: datetime = datetime.date(self.creation_date.year, self.creation_date.month - 1, 1)
        self.end_billing_period: datetime = self._get_last_day_of_month(self.creation_date.year,
                                                                        self.creation_date.month)

        self.workbook: Workbook = Workbook()
        self.sheet: Worksheet = Worksheet('')

        self.last_row = {"days": 0,
                         "hours": 0
                         }

        self.worker_counter = 0

        self.__assemble_workbook()

    @logger.catch
    def _merge(self, cells, cell, text, alignment=Alignment(horizontal='center', vertical='center', wrap_text=True),
               font=Font(name='Arial', size=9)):
        self.sheet.merge_cells(cells)
        self.sheet[cell].value = text
        self.sheet[cell].alignment = alignment
        self.sheet[cell].font = font

    @logger.catch
    def _merge_wrap_text(self, cells, cell, text):
        self.sheet.merge_cells(cells)
        self.sheet[cell].value = text
        self.sheet[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _get_last_day_of_month(year, month):
        num_days = calendar.monthrange(year, month - 1)[1]
        return datetime.date(year, month - 1, num_days)

    @staticmethod
    def _get_days_with_eights(year: int, month: int, working_hours,
                              employment_date: datetime, dismissal_date: datetime, not_x=False):
        # Определяем количество дней в месяце
        num_days = calendar.monthrange(year, month)[1]

        def _until_this_day(day: int, hiring_day: int, _month: int, hiring_month: int) -> bool:
            if year == employment_date.year and month == hiring_month and day < hiring_day:
                return False
            if year == employment_date.year and month < hiring_month:
                return False
            if year == dismissal_date.year and month > dismissal_date.month:
                return False
            if year == dismissal_date.year and month == dismissal_date.month and day >= dismissal_date.day:
                return False
            return True

        if not_x:
            days_data = (int(working_hours) if day.weekday() < 5 and _until_this_day(day.day, employment_date.day,
                                                                                     month, employment_date.month) else 0 for day in
                         (datetime.date(year, month, day_num) for day_num in range(1, num_days + 1)))
            return tuple(days_data)

        # Создаем кортеж для хранения данных
        days_data = (f'{working_hours}' if day.weekday() < 5 and _until_this_day(day.day, employment_date.day, month, employment_date.month) else 'x' for day in (datetime.date(year, month, day_num) for day_num in range(1, num_days + 1)))

        return tuple(days_data)

    def _days_in_month_header_block(self):
        days_in_month = self.end_billing_period.day

        # ROW I
        self._merge('G13:G14', 'G13', '01')
        self._merge('H13:H14', 'H13', '02')
        self._merge('I13:I14', 'I13', '03')
        self._merge('J13:J14', 'J13', '04')
        self._merge('K13:K14', 'K13', '05')
        self._merge('L13:L14', 'L13', '06')
        self._merge('M13:M14', 'M13', '07')
        self._merge('N13:N14', 'N13', '08')
        self._merge('O13:O14', 'O13', '09')
        self._merge('P13:P14', 'P13', '10')
        self._merge('Q13:Q14', 'Q13', '11')
        self._merge('R13:R14', 'R13', '12')
        self._merge('S13:S14', 'S13', '13')
        self._merge('T13:T14', 'T13', '14')
        self._merge('U13:U14', 'U13', '15')
        self._merge('V13:V14', 'V13', 'x')

        # RPOW II
        self._merge('G15:G17', 'G15', '16')
        self._merge('H15:H17', 'H15', '17')
        self._merge('I15:I17', 'I15', '18')
        self._merge('J15:J17', 'J15', '19')
        self._merge('K15:K17', 'K15', '20')
        self._merge('L15:L17', 'L15', '21')
        self._merge('M15:M17', 'M15', '22')
        self._merge('N15:N17', 'N15', '23')
        self._merge('O15:O17', 'O15', '24')
        self._merge('P15:P17', 'P15', '25')
        self._merge('Q15:Q17', 'Q15', '26')
        self._merge('R15:R17', 'R15', '27')
        self._merge('S15:S17', 'S15', '28')
        self._merge('T15:T17', 'T15', '29' if days_in_month >= 29 else 'x')
        self._merge('U15:U17', 'U15', '30' if days_in_month >= 30 else 'x')
        self._merge('V15:V17', 'V15', '31' if days_in_month == 31 else 'x')

    @logger.catch
    def _sum_days(self, days: tuple[str]) -> int:
        result = 0
        for day in days:
            if day != 'x':
                result += 1
        return result

    def _sum_hours(self, days: tuple[str]) -> int:
        result = 0
        for day in days:
            if day != 'x':
                result += int(day)
        return result

    def _if_employment_later_last_month(self, worker: Worker):
        today = datetime.date.today()
        first_day_of_this_month = today.replace(day=1)
        last_day_of_last_month = first_day_of_this_month - datetime.timedelta(days=1)
        previous_month = last_day_of_last_month.month

        if int(worker.employment_date.split('.')[1]) > previous_month:
            return True

        if worker.dismissal != '':
            if int(worker.dismissal.split('.')[1]) < previous_month:
                return True

        return False

    def _fill_table(self, num, worker: Worker, row_num: int):
        if worker.if_employment_later_last_month():
            return True

        self.worker_counter += 1

        last_row = row_num

        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        bottom_line_border = Border(
            bottom=Side(border_style='thin', color='000000')
        )

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        employment_date = datetime.datetime.strptime(worker.employment_date, '%d.%m.%Y')

        if worker.dismissal == '':
            dismissal_date = datetime.datetime.strptime('10.10.2099', '%d.%m.%Y')
        else:
            dismissal_date = datetime.datetime.strptime(worker.dismissal, '%d.%m.%Y')
        try:
            days = self._get_days_with_eights(self.start_billing_period.year, self.start_billing_period.month,
                                              worker.working_hours, employment_date,
                                              dismissal_date)
        except ValueError:
            raise WorkerNotHaveWorkHours(worker)

        self._merge(f'A{last_row}:A{last_row + 1}', f'A{last_row}', f'{num + 1}')
        self._merge(f'B{last_row}:B{last_row + 1}', f'B{last_row}', f'{num + 1}')
        self._merge(f'C{last_row}:C{last_row + 1}', f'C{last_row}', worker.sex)

        if '.' in worker.name:
            self._merge_wrap_text(f'D{last_row}:F{last_row + 1}', f'D{last_row}', worker.name)
        else:
            try:
                self._merge_wrap_text(f'D{last_row}:F{last_row + 1}', f'D{last_row}',
                                      f'{worker.name.split(" ")[0]}\n{worker.name.split(" ")[1]} {worker.name.split(" ")[2]}')
            except IndexError:
                self._merge_wrap_text(f'D{last_row}:F{last_row + 1}', f'D{last_row}', worker.name)

        self.sheet[f'G{last_row}'] = days[0]
        self.sheet[f'H{last_row}'] = days[1]
        self.sheet[f'I{last_row}'] = days[2]
        self.sheet[f'J{last_row}'] = days[3]
        self.sheet[f'K{last_row}'] = days[4]
        self.sheet[f'L{last_row}'] = days[5]
        self.sheet[f'M{last_row}'] = days[6]
        self.sheet[f'N{last_row}'] = days[7]
        self.sheet[f'O{last_row}'] = days[8]
        self.sheet[f'P{last_row}'] = days[9]
        self.sheet[f'Q{last_row}'] = days[10]
        self.sheet[f'R{last_row}'] = days[11]
        self.sheet[f'S{last_row}'] = days[12]
        self.sheet[f'T{last_row}'] = days[13]
        self.sheet[f'U{last_row}'] = days[14]
        self.sheet[f'V{last_row}'] = 'x'
        self._merge(f'W{last_row}:W{last_row + 1}', f'W{last_row}', f'{self._sum_days(days)}')
        try:
            self.last_row["days"] += self._sum_days(days)
            self._merge(f'X{last_row}:X{last_row + 1}', f'X{last_row}', f"{self._sum_hours(days)}")

            self.last_row["hours"] += self._sum_hours(days)
            self._merge(f'AP{last_row}:AP{last_row + 1}', f'AP{last_row}', worker.salary_real)
        except TypeError:
            raise NoWorkers()
        except ValueError:
            raise WorkerNotHaveWorkHours(worker)

        for row in self.sheet.iter_rows(min_row=last_row - 1, max_row=last_row, min_col=1, max_col=42):
            for cell in row:
                cell.border = border
                cell.alignment = center_alignment

        last_row += 1

        self.sheet[f'G{last_row}'] = days[15]
        self.sheet[f'H{last_row}'] = days[16]
        self.sheet[f'I{last_row}'] = days[17]
        self.sheet[f'J{last_row}'] = days[18]
        self.sheet[f'K{last_row}'] = days[19]
        self.sheet[f'L{last_row}'] = days[20]
        self.sheet[f'M{last_row}'] = days[21]
        self.sheet[f'N{last_row}'] = days[22]
        self.sheet[f'O{last_row}'] = days[23]
        self.sheet[f'P{last_row}'] = days[24]
        self.sheet[f'Q{last_row}'] = days[25]
        self.sheet[f'R{last_row}'] = days[26]
        self.sheet[f'S{last_row}'] = days[27]
        if len(days) >= 29:
            self.sheet[f'T{last_row}'] = days[28]
        else:
            self.sheet[f'T{last_row}'] = 'x'
        if len(days) >= 30:
            self.sheet[f'U{last_row}'] = days[29]
        else:
            self.sheet[f'U{last_row}'] = 'x'
        if len(days) == 31:
            self.sheet[f'V{last_row}'] = days[30]
        else:
            self.sheet[f'V{last_row}'] = 'x'

        for row in self.sheet.iter_rows(min_row=last_row, max_row=last_row, min_col=1, max_col=42):
            for cell in row:
                cell.border = border
                cell.alignment = center_alignment
        self.sheet.row_dimensions[last_row - 1].height = 25
        self.sheet.row_dimensions[last_row].height = 25

        return True

    def _fill_last_row(self, start_row: int):
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        self._merge(f'A{start_row}:V{start_row}', f'A{start_row}', 'РАЗОМ:',
                    alignment=Alignment(horizontal='left', vertical='center', wrap_text=True),
                    font=Font(name='Arial', size=9, bold=True))
        self.sheet[f'W{start_row}'] = self.last_row['days']
        self.sheet[f'X{start_row}'] = self.last_row['hours']

        font = Font(name='Arial', bold=False, italic=False, size=8)

        for row in self.sheet.iter_rows(min_row=start_row, max_row=start_row, min_col=1, max_col=42):
            for cell in row:
                cell.border = border
                cell.alignment = center_alignment
                cell.font = font

        self._merge(f'A{start_row}:V{start_row}', f'A{start_row}', 'РАЗОМ:',
                    alignment=Alignment(horizontal='right', vertical='center', wrap_text=True),
                    font=Font(name='Arial', size=9, bold=True))

    def __assemble_workbook(self):

        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        bottom_line_border = Border(
            bottom=Side(border_style='thin', color='000000')
        )

        self.sheet: Worksheet = self.workbook.active

        self._merge('AM1:AP3', 'AM1',
                    'Типова форма № П-5\nЗАТВЕРДЖЕНО\nНаказ Держкомстату України\n05.12.2008  № 489',
                    alignment=Alignment(horizontal='left', vertical='center', ))

        self._merge('B2:AI2', 'B2', self.Employer.name)

        self._merge('B3:AI3', 'B3', 'Найменування  підприємства (установи, організації)')

        self._merge('B5:AI5', 'B5', '')

        self._merge('B6:AI6', 'B6', 'назва структурного підрозділу')

        self._merge('AJ7:AK7', 'AJ7', 'Дата заповнення')
        self._merge('AL7:AO7', 'AL7', 'Звітний період')

        self.sheet['B8'] = 'Ідентифікаційний код  ЄДРПОУ'
        self.sheet['B8'].font = Font(name='Arial', bold=False, size=10)

        self._merge('G8:U8', 'G8', self.Employer.ident_EDRPOU)
        self._merge('AJ8:AK9', 'AJ8', self.start_writing_period.strftime('%d.%m.%Y'))
        self._merge('AL8:AM8', 'AL8', 'з')
        self._merge('AN8:AO8', 'AN8', 'по')

        self._merge('AL9:AM9', 'AL9', self.start_billing_period.strftime('%d.%m.%Y'))
        self._merge('AN9:AO9', 'AN9', self.end_billing_period.strftime('%d.%m.%Y'))

        self._merge('B10:AI10', 'B10', 'ТАБЕЛЬ ОБЛІКУ ВИКОРИСТАННЯ РОБОЧОГО ЧАСУ')

        self._merge('A12:A17', 'A12', '№  П/П')
        self._merge('B12:B17', 'B12', 'Табельний  номер')
        self._merge('C12:C17', 'C12', 'Стать  (ч/ж)')
        self._merge('D12:F17', 'D12', 'ПІБ, посада')
        self._merge('G12:V12', 'G12', 'Відмітки  про явки та неявки за числами місяця (годин)')
        self._merge('W12:AB12', 'W12', 'Відпрацьовано за місяць')
        self._merge('AC12:AC14', 'AC12', 'Всього\nнеявок')
        self._merge('AD12:AO12', 'AD12', 'з причин за місяць')
        self._merge_wrap_text('AP12:AP17', 'AP12', 'Оклад, тарифна ставка, грн')

        self._days_in_month_header_block()

        self._merge('W13:W17', 'W13', 'днів')
        self._merge('X13:AB13', 'X13', 'годин')
        self._merge_wrap_text('AD13:AD14', 'AD13', 'основна та додаткова відпустки')
        self._merge_wrap_text('AE13:AE14', 'AE13', 'відпустки у зв’язку з навчанням, творчі, в обов. порядку та інші')
        self._merge_wrap_text('AF13:AF14', 'AF13', 'відпустки без збереження заробітної плати за згодою сторін')
        self._merge_wrap_text('AG13:AG14', 'AG13', 'відпустки без збереження з/п на період припинення виконання робіт')
        self._merge_wrap_text('AH13:AH14', 'AH13', 'перевод на неповний робочий день (тиждень)')
        self._merge_wrap_text('AI13:AI14', 'AI13', 'тимчасовий перевод на інше підприємство')
        self._merge_wrap_text('AJ13:AJ14', 'AJ13', 'простої')
        self._merge_wrap_text('AK13:AK14', 'AK13', 'прогули')
        self._merge_wrap_text('AL13:AL14', 'AL13', 'страйки')
        self._merge_wrap_text('AM13:AM14', 'AM13', 'тимчасова непрацездатність')
        self._merge_wrap_text('AN13:AN14', 'AN13', 'інші')
        self._merge_wrap_text('AO13:AO14', 'AO13', '')

        self._merge('X14:X17', 'X14', 'всього')
        self._merge('Y14:AB14', 'Y14', 'з них:')

        self._merge_wrap_text('Y15:Y17', 'Y15', 'над- уроч- но')
        self._merge_wrap_text('Z15:Z17', 'Z15', 'ніч- них')
        self._merge_wrap_text('AA15:AA17', 'AA15', 'вечір-ніх')
        self._merge_wrap_text('AB15:AB17', 'AB15', 'вихідних, святко- вих')
        self._merge('AC15:AC16', 'AC15', 'години')
        self._merge_wrap_text('AD15:AD16', 'AD15', 'коди\n8-10')
        self._merge_wrap_text('AE15:AE16', 'AE15', 'коди\n11-15, 17,22')
        self._merge_wrap_text('AF15:AF16', 'AF15', 'коди\n18')
        self._merge_wrap_text('AG15:AG16', 'AG15', 'коди\n19')
        self._merge_wrap_text('AH15:AH16', 'AH15', 'коди\n20')
        self._merge_wrap_text('AI15:AI16', 'AI15', 'коди\n21')
        self._merge_wrap_text('AJ15:AJ16', 'AJ15', 'коди\n23')
        self._merge_wrap_text('AK15:AK16', 'AK15', 'коди\n24')
        self._merge_wrap_text('AL15:AL16', 'AL15', 'коди\n25')
        self._merge_wrap_text('AM15:AM16', 'AM15', 'коди\n26-27')
        self._merge_wrap_text('AN15:AN16', 'AN15', 'коди\n28-30')
        self._merge_wrap_text('AO15:AO16', 'AO15', '')

        self.sheet['AC17'] = 'дні'
        self.sheet['AD17'] = 'дні/год.'
        self.sheet['AE17'] = 'дні/год.'
        self.sheet['AF17'] = 'дні/год.'
        self.sheet['AG17'] = 'дні/год.'
        self.sheet['AH17'] = 'дні/год.'
        self.sheet['AI17'] = 'дні/год.'
        self.sheet['AJ17'] = 'дні/год.'
        self.sheet['AK17'] = 'дні/год.'
        self.sheet['AL17'] = 'дні/год.'
        self.sheet['AM17'] = 'дні/год.'
        self.sheet['AN17'] = 'дні/год.'
        self.sheet['AO17'] = 'дні/год.'

        last_row = 18

        # Заполняем таблицу
        for num, worker in enumerate(self.Employer.workers):
            if_feel = self._fill_table(num, worker, last_row)
            if if_feel:
                last_row += 2

        font = Font(name='Arial', bold=False, italic=False, size=8)
        for col in self.sheet.columns:
            for cell in col:
                cell.font = font

        if self.worker_counter == 0:
            raise NoWorkers()

        self._fill_last_row(last_row)

        last_row += 2

        self.sheet[f'B{last_row}'].font = Font(name='Arial', bold=True, size=9)
        self.sheet[f'B{last_row}'] = 'Відповідальна особа'

        self._merge(f'F{last_row}:Q{last_row}', f'F{last_row}', 'Директор')
        self.sheet[f'F{last_row}'].font = Font(name='Arial', bold=True, size=9)

        self.sheet[f'U{last_row}'] = 'Керівник структурного'
        self.sheet[f'U{last_row}'].font = Font(name='Arial', bold=True, size=9)

        self._merge(f'Z{last_row}:AK{last_row}', f'Z{last_row}', 'Директор')
        self.sheet[f'Z{last_row}'].font = Font(name='Arial', bold=True, size=9)

        last_row += 1

        self._merge(f'F{last_row}:Q{last_row}', f'F{last_row}', '(посада)')
        self.sheet[f'U{last_row}'] = 'підрозділу'
        self.sheet[f'U{last_row}'].font = Font(bold=True)
        self._merge(f'Z{last_row}:AF{last_row}', f'Z{last_row}', '(посада)')
        self.sheet[f'U{last_row}'].font = Font(name='Arial', bold=True, size=9)

        last_row += 3

        font = Font(name='Arial', bold=False, italic=False, size=8)

        for row in self.sheet.iter_rows(min_row=last_row, max_row=last_row+3, min_col=1, max_col=42):
            for cell in row:
                cell.font = font
                cell.alignment = Alignment(horizontal='left', vertical='bottom', )

        self._merge(f'F{last_row}:H{last_row}', f'F{last_row}', '')
        self._merge(f'B{last_row}:E{last_row}', f'B{last_row}', '\"____\"___________20___р.')
        self.sheet[f'J{last_row}'] = self.Employer.name
        self._merge(f'U{last_row}:Y{last_row}', f'U{last_row}', '\"____\"___________20___р.')
        self._merge(f'Z{last_row}:AB{last_row}', f'Z{last_row}', '')
        self.sheet[f'AD{last_row}'] = self.Employer.name

        last_row += 1

        self._merge(f'F{last_row}:H{last_row}', f'F{last_row}', '(підпис)')
        self._merge(f'J{last_row}:Q{last_row}', f'J{last_row}', '(ПІБ)')
        self._merge(f'Z{last_row}:AB{last_row}', f'Z{last_row}', '(підпис)')
        self._merge(f'AD{last_row}:AF{last_row}', f'AD{last_row}', '(ПІБ)')

        self.sheet['B2'].font = Font(name='Arial', bold=True, size=10)
        self.sheet['G8'].font = Font(name='Arial', bold=True, size=10)
        self.sheet['AJ8'].font = Font(name='Arial', bold=True, size=9)
        self.sheet['AL9'].font = Font(name='Arial', bold=True, size=9)
        self.sheet['AN9'].font = Font(name='Arial', bold=True, size=9)
        self.sheet['B10'].font = Font(name='Arial', bold=True, size=14)
        cell = self.sheet['A12']
        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        cell = self.sheet['B12']
        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        cell = self.sheet['C12']
        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        cell = self.sheet['W13']
        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        cell = self.sheet['X14']
        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

        self.sheet.column_dimensions['A'].width = 3.46
        self.sheet.column_dimensions['B'].width = 7.26
        self.sheet.column_dimensions['C'].width = 3.2
        self.sheet.column_dimensions['D'].width = 6.5
        self.sheet.column_dimensions['E'].width = 6.5
        self.sheet.column_dimensions['F'].width = 8.77
        self.sheet.column_dimensions['G'].width = 3.2
        self.sheet.column_dimensions['H'].width = 3.2
        self.sheet.column_dimensions['I'].width = 3.2
        self.sheet.column_dimensions['J'].width = 3.2
        self.sheet.column_dimensions['K'].width = 3.2
        self.sheet.column_dimensions['L'].width = 3.2
        self.sheet.column_dimensions['M'].width = 3.2
        self.sheet.column_dimensions['N'].width = 3.2
        self.sheet.column_dimensions['O'].width = 3.2
        self.sheet.column_dimensions['P'].width = 3.2
        self.sheet.column_dimensions['Q'].width = 3.2
        self.sheet.column_dimensions['R'].width = 3.2
        self.sheet.column_dimensions['S'].width = 3.2
        self.sheet.column_dimensions['T'].width = 3.2
        self.sheet.column_dimensions['U'].width = 3.2
        self.sheet.column_dimensions['V'].width = 3.2
        self.sheet.column_dimensions['W'].width = 6.47
        self.sheet.column_dimensions['X'].width = 6.47
        self.sheet.column_dimensions['Y'].width = 6.47
        self.sheet.column_dimensions['Z'].width = 6.47
        self.sheet.column_dimensions['AA'].width = 6.47
        self.sheet.column_dimensions['AB'].width = 6.47
        self.sheet.column_dimensions['AC'].width = 6.47
        self.sheet.column_dimensions['AD'].width = 6.47
        self.sheet.column_dimensions['AE'].width = 6.47
        self.sheet.column_dimensions['AF'].width = 6.47
        self.sheet.column_dimensions['AG'].width = 6.47
        self.sheet.column_dimensions['AH'].width = 6.47
        self.sheet.column_dimensions['AI'].width = 6.47
        self.sheet.column_dimensions['AJ'].width = 6.47
        self.sheet.column_dimensions['AK'].width = 6.47
        self.sheet.column_dimensions['AL'].width = 6.47
        self.sheet.column_dimensions['AM'].width = 6.47
        self.sheet.column_dimensions['AN'].width = 6.47
        self.sheet.column_dimensions['AO'].width = 6.47
        self.sheet.column_dimensions['AP'].width = 8.33

        self.sheet.row_dimensions[1].height = 14
        self.sheet.row_dimensions[2].height = 12.75
        self.sheet.row_dimensions[3].height = 14
        self.sheet.row_dimensions[4].height = 6
        self.sheet.row_dimensions[5].height = 12.75
        self.sheet.row_dimensions[6].height = 12.75
        self.sheet.row_dimensions[7].height = 12.75
        self.sheet.row_dimensions[8].height = 12.75
        self.sheet.row_dimensions[9].height = 12.75
        self.sheet.row_dimensions[10].height = 18
        self.sheet.row_dimensions[11].height = 12.75
        self.sheet.row_dimensions[12].height = 12.5
        self.sheet.row_dimensions[13].height = 53
        self.sheet.row_dimensions[14].height = 88.5
        self.sheet.row_dimensions[15].height = 16.25
        self.sheet.row_dimensions[16].height = 17.25
        self.sheet.row_dimensions[17].height = 14.75
        # self.sheet.row_dimensions[18].height = 11.25
        # self.sheet.row_dimensions[19].height = 11.25
        # self.sheet.row_dimensions[20].height = 12.75
        # self.sheet.row_dimensions[21].height = 14
        # self.sheet.row_dimensions[22].height = 14
        # self.sheet.row_dimensions[23].height = 12.75
        # self.sheet.row_dimensions[24].height = 12.75
        # self.sheet.row_dimensions[25].height = 11.25
        # self.sheet.row_dimensions[26].height = 11.25
        # self.sheet.row_dimensions[27].height = 12.75

        for row in self.sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=35):
            for cell in row:
                cell.border = bottom_line_border

        for row in self.sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=35):
            for cell in row:
                cell.border = bottom_line_border

        for row in self.sheet.iter_rows(min_row=7, max_row=9, min_col=36, max_col=41):
            for cell in row:
                cell.border = border

        for row in self.sheet.iter_rows(min_row=8, max_row=8, min_col=7, max_col=21):
            for cell in row:
                cell.border = bottom_line_border

        for row in self.sheet.iter_rows(min_row=12, max_row=19, min_col=1, max_col=42):
            for cell in row:
                cell.border = border

        # Director row line
        for row in self.sheet.iter_rows(min_row=last_row - 5, max_row=last_row - 5, min_col=6, max_col=17):
            for cell in row:
                cell.border = bottom_line_border
        for row in self.sheet.iter_rows(min_row=last_row - 5, max_row=last_row - 5, min_col=26, max_col=26):
            for cell in row:
                cell.border = bottom_line_border
        # Last row linews
        for row in self.sheet.iter_rows(min_row=last_row - 1, max_row=last_row - 1, min_col=6, max_col=17):
            for cell in row:
                cell.border = bottom_line_border

        for row in self.sheet.iter_rows(min_row=last_row - 1, max_row=last_row - 1, min_col=26, max_col=37):
            for cell in row:
                cell.border = bottom_line_border

        for row in self.sheet.iter_rows(min_row=last_row - 1, max_row=last_row - 1, min_col=10, max_col=17):
            for cell in row:
                cell.border = bottom_line_border

        # self.sheet['F25'].border = bottom_line_border
        # self.sheet['G25'].border = bottom_line_border
        # self.sheet['H25'].border = bottom_line_border
        # self.sheet['Z25'].border = bottom_line_border
        # self.sheet['AA25'].border = bottom_line_border
        # self.sheet['AB25'].border = bottom_line_border
        # self.sheet['AD25'].border = bottom_line_border
        # self.sheet['AE25'].border = bottom_line_border
        # self.sheet['AF25'].border = bottom_line_border

    def get_bytes(self) -> io.BytesIO:
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def save(self):
        self.workbook.save("example.xlsx")


if __name__ == '__main__':
    worker1 = Worker(
        sex="М",
        name="John Doe",
        job_title="Software Engineer",
        salary="5000",
        working_hours="8",
        ident_IPN="1234567890",
        employment_date="15.09.2023",
        birthday="13.03.1992",
        dismissal=''
    )

    worker2 = Worker(
        sex="Ж",
        name="Jane Smith",
        job_title="Data Scientist",
        salary="6000",
        working_hours="8",
        ident_IPN="0987654321",
        employment_date="18.07.2023",
        birthday="13.03.1992",
        dismissal=''
    )

    # Создание объекта Employer с несколькими работниками
    employer = Employer(
        name="Acme Corporation",
        ident_EDRPOU="123456789",
        workers=[worker1, worker2],
        residence='UK',
        phone='09094594095'
    )

    sheet = AppearanceOTWHSheet(employer)
    sheet.save()
