import calendar
import datetime
import io

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles.borders import BORDER_MEDIUM
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, Side

from Exceptions import WorkerNotHaveWorkHours
from logger import Logger
from tables.Models import Employer, Worker
from XLAssembler import months_names
from .working_hour_sheet import AppearanceOTWHSheet

cls_logger = Logger()
logger = cls_logger.get_logger


class SettlementPayment:
    def __init__(self, employer: Employer):

        self.Employer: Employer = employer
        self.creation_date: datetime = datetime.datetime.now()
        self.start_billing_period: datetime = datetime.date(self.creation_date.year, self.creation_date.month, 1) - relativedelta(months=1)
        self.end_billing_period: datetime = AppearanceOTWHSheet._get_last_day_of_month(self.creation_date.year,
                                                                                       self.creation_date.month) - relativedelta(months=1)
        self.worker_counter = 0

        self.workbook: Workbook = Workbook()
        self.sheet: Worksheet = Worksheet('')

        self.last_row = {"days": 0,
                         "F": 0.0,
                         "P": 0.0,
                         "R": 0.0,
                         "U": 0.0,
                         "V": 0.0,
                         "Y": 0.0,
                         "Z": 0.0,
                         "AA": 0.0,
                         "AC": 0.0
                         }

        self.__assemble_workbook()

    def _merge(self, cells, cell, text, font=None):

        self.sheet.merge_cells(cells)
        top_left_cell = self.sheet[cell]
        top_left_cell.value = text
        top_left_cell.alignment = Alignment(horizontal='center', vertical='center')
        if font:
            top_left_cell.font = font

    def _billing_period(self, get_int: bool = False) -> str:
        def get_actual_month(_month: int) -> int:
            if _month == 1:
                return 12
            return _month - 1

        month = months_names[get_actual_month(self.creation_date.month)-1]
        year = self.creation_date.year if month != 'грудень' else self.creation_date.year - 1

        if get_int:
            return f'{get_actual_month(self.creation_date.month)} {year}'

        return f'{month} {year}'

    def _if_employment_later_last_month(self, worker: Worker):
        today = datetime.date.today()
        first_day_of_this_month = today.replace(day=1)
        last_day_of_last_month = first_day_of_this_month - datetime.timedelta(days=1)
        previous_month = last_day_of_last_month.month

        if int(worker.employment_date.split('.')[1]) > previous_month:
            return True

        if worker.dismissal != '':
            if int(worker.dismissal.split('.')[1]) < previous_month:
                return True

        return False

    def _get_maximum_work_days(self, worker: Worker) -> int:
        result = 0
        _month = int(self._billing_period(get_int=True).split(' ')[0])
        _year = int(self._billing_period().split(' ')[1])
        num_days = calendar.monthrange(_year, _month)[1]

        days_data = (int(worker.working_hours) if day.weekday() < 5 else 0 for day in
                         (datetime.date(self.start_billing_period.year,
                          self.start_billing_period.month, day_num) for day_num in range(1, num_days + 1)))

        for day in days_data:
            if day != 0:
                result += 1

        return result

    def _fill_table(self, num: int, worker: Worker, start_row: int):
        if worker.if_employment_later_last_month():
            return False

        _start_row = start_row

        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        employment_date = datetime.datetime.strptime(worker.employment_date, '%d.%m.%Y')
        if worker.dismissal == '':
            dismissal_date = datetime.datetime.strptime('10.10.2099', '%d.%m.%Y')
        else:
            dismissal_date = datetime.datetime.strptime(worker.dismissal, '%d.%m.%Y')

        try:
            work_days = int(sum(AppearanceOTWHSheet._get_days_with_eights(self.start_billing_period.year,
                                                                          self.start_billing_period.month,
                                                                          worker.working_hours, employment_date,
                                                                          dismissal_date, not_x=True)) / int(
                worker.working_hours))
        except ValueError:
            raise WorkerNotHaveWorkHours(worker)

        if work_days < 1:
            return False

        self.worker_counter += 1

        self.sheet[f'A{_start_row}'] = f'{self.worker_counter}'
        self.sheet[f'B{_start_row}'] = f'{self.worker_counter}'
        self.sheet[f'D{_start_row}'] = f'{worker.job_title}'

        self.last_row["days"] += work_days

        self.sheet[f'E{_start_row}'] = f'{work_days}'

        max_days = self._get_maximum_work_days(worker)
        _salary_per_day = int(worker.salary_real)/max_days
        _salary = round(_salary_per_day * work_days, 2)

        self.last_row["F"] += _salary
        self.sheet[f'F{_start_row}'] = f'{_salary}'

        for cells in self.sheet.iter_cols(min_col=1, max_col=self.sheet.max_column,
                                          min_row=_start_row, max_row=_start_row):
            for cell in cells:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

        start_column = 7

        for column in range(start_column, self.sheet.max_column-3):
            self.sheet.cell(row=_start_row, column=column).value = '-'

        self.last_row["P"] += _salary
        self.sheet[f'P{_start_row}'] = f'{_salary}'

        q21 = round(float(_salary) * 0.18, 2)
        self.last_row["R"] += q21
        self.sheet[f'R{_start_row}'] = f"{q21}"

        t21 = round(float(_salary) * 0.015, 2)
        self.last_row["U"] += t21
        self.sheet[f'U{_start_row}'] = f"{t21}"

        u21 = q21 + t21
        self.last_row["V"] += round(u21, 2)
        self.sheet[f'V{_start_row}'] = f"{round(u21, 2)}"

        x21 = round(float(_salary) - u21, 2)
        self.last_row["Y"] += round(x21, 2)
        self.sheet[f'Y{_start_row}'] = f"{x21}"

        self.sheet[f'Z{_start_row}'] = f'{x21}'
        self.last_row["Z"] += x21

        self.sheet[f'AA{_start_row}'] = f'{x21}'
        self.last_row["AA"] += x21

        self.sheet[f'AB{_start_row}'] = ''
        self.sheet[f'AC{_start_row}'] = f'{worker.name} {worker.ident_IPN}'

        self.sheet.row_dimensions[_start_row].height = 46

        def replace_n(string: str) -> str:
            return string.replace('\n', '').replace('\t', '') #.replace(' ', '')

        name = worker.name.replace(' - ', '-')

        if '.' in name:
            self.sheet[
                f'AC{_start_row}'] = f'{replace_n(name)}\n{replace_n(worker.ident_IPN)}'
        else:
            try:
                self.sheet[f'AC{_start_row}'] = f'{name.split(" ")[0]}\n{name.split(" ")[1]} {replace_n(name.split(" ")[2])} {replace_n(worker.ident_IPN)}'
            except IndexError:
                self.sheet[
                    f'AC{_start_row}'] = f'{replace_n(name)}\n{replace_n(worker.ident_IPN)}'

        return True

    def _fill_last_row(self, start_row: int):
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        self._merge(f'A{start_row}:B{start_row}', f'A{start_row}', 'Разом:')
        self.sheet[f'E{start_row}'] = f'{self.last_row["days"]}'
        self.sheet[f'F{start_row}'] = f'{round(self.last_row["F"], 2)}'
        self.sheet[f'P{start_row}'] = f'{round(self.last_row["P"], 2)}'
        self.sheet[f'R{start_row}'] = f'{round(self.last_row["R"], 2)}'
        self.sheet[f'U{start_row}'] = f'{round(self.last_row["U"], 2)}'
        self.sheet[f'V{start_row}'] = f'{round(self.last_row["V"], 2)}'

        self.sheet[f'Y{start_row}'] = f'{round(self.last_row["Y"], 2)}'
        self.sheet[f'Z{start_row}'] = f'{round(self.last_row["Z"], 2)}'
        self.sheet[f'AA{start_row}'] = f'{round(self.last_row["AA"], 2)}'

        for cells in self.sheet.iter_cols(min_col=1, max_col=self.sheet.max_column,
                                          min_row=start_row, max_row=start_row):
            for cell in cells:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

    def __assemble_workbook(self):
        self.sheet = self.workbook.active

        self.sheet['A1'] = f'ФОП {self.Employer.name}'

        self._merge('A2:D2', 'A2', 'Ідентифікаційний код за ЄДРПОУ')
        self._merge('E2:F2', 'E2', f'{self.Employer.ident_EDRPOU}', font=Font(name='Arial', bold=True, size=9))
        self.sheet['E2'].font = Font(bold=True)

        self.sheet['B3'] = 'Цех, відділ'

        self.sheet['C5'] = 'В касу для оплати в строк з _________________________________ '
        self.sheet['H5'] = 'до  ________________________________'
        self.sheet['L5'] = 'року'

        self.sheet['C6'] = 'В сумі'

        self.sheet['C7'] = f'Керівник   {self.Employer.name}'

        self.sheet['C8'] = f'Головний бухгалтер   {self.Employer.name}'

        settle_num = datetime.datetime.now().month-1
        settle_num = settle_num if settle_num != 0 else 12
        self._merge('B10:O10', 'B10',
                    f'Розрахунково-платіжна відомість № {settle_num} за {self._billing_period()} р.')
        top_left_cell = self.sheet['B10']
        top_left_cell.font = Font(bold=True)

        self._merge('F12:O12', 'F12', 'Нараховано за видами виплат')
        self._merge('P12:U12', 'P12', 'Утримано та враховано')
        self._merge('X12:Z12', 'X12', 'Сума, грн.')

        self._merge('M13:O13', 'M13', 'Допомога за')
        self.sheet['R13'] = 'Податок\nна\nдоходи\nфіз.осіб'
        self.sheet['R15'] = ''
        self.sheet['R16'] = ''
        self.sheet['R17'] = ''

        self.sheet['E14'] = 'Відпра-'
        self._merge('F14:J14', 'F14', 'Доплат та надбавок')

        self.sheet['L14'] = 'За'
        self.sheet['L15'] = 'чергову'
        self.sheet['L16'] = 'відпустку'

        self._merge('M14:O14', 'M14', 'тимчасовою')
        self.sheet['S14'] = 'Єдиний'
        self.sheet['T14'] = 'Проф-'
        self.sheet['W14'] = 'Заборго-'
        self.sheet['X14'] = 'Виплачено'

        self.sheet['A15'] = '№'
        self.sheet['A15'].alignment = Alignment(horizontal='center', vertical='center')
        self.sheet['B15'] = 'Табельний'
        self.sheet['C15'] = 'Категорія'
        self.sheet['D15'] = 'Професія,'
        self.sheet['E15'] = 'цьовано'
        self._merge('M15:O15', 'M15', 'непрацездатністю')
        self.sheet['Q15'] = 'виплачено'
        self.sheet['S15'] = 'соціаль-'
        self.sheet['T15'] = 'спілко-'
        self.sheet['U15'] = 'Інші'
        self.sheet['W15'] = 'Разом'
        self.sheet['X15'] = 'за'
        self.sheet['Y15'] = 'Заробіт-'
        self.sheet['Z15'] = 'Належить'
        self.sheet['AA15'] = 'Разом'
        self.sheet['AB15'] = 'Розписка'
        self.sheet['AC15'] = 'Прізвище,'

        self.sheet['A16'] = 'з/п'
        self.sheet['B16'] = 'номер'
        self.sheet['C16'] = 'персоналу'
        self.sheet['D16'] = 'посада,'
        self.sheet['E16'] = 'днів,'
        self.sheet['F16'] = 'За тариф.'
        self.sheet['G16'] = 'Благодійна'
        self.sheet['H16'] = 'Обклад.'
        self.sheet['I16'] = 'Обклад.'
        self.sheet['J16'] = 'Індек-'
        self.sheet['K16'] = 'Премія'
        self.sheet['P16'] = 'Разом,'
        self.sheet['Q16'] = 'аванса'
        self.sheet['S16'] = 'ний'
        self.sheet['T16'] = 'вий'
        self.sheet['U16'] = 'утри-'
        self.sheet['V16'] = 'Разом'
        self.sheet['W16'] = 'за/перед'
        self.sheet['X16'] = 'попередні'
        self.sheet['Y16'] = 'ної'
        self.sheet['Z16'] = 'до'
        self.sheet['AA16'] = 'до'
        self.sheet['AB16'] = 'в'
        self.sheet['AC16'] = 'ім\'я,'

        self.sheet['E17'] = 'годин'
        self.sheet['F17'] = 'ставками'
        self.sheet['G17'] = 'допомога'
        self.sheet['H17'] = 'податками'
        self.sheet['I17'] = 'податками'
        self.sheet['J17'] = 'сація'
        self.sheet['M17'] = 'місяць'
        self.sheet['N17'] = 'дні'
        self.sheet['O17'] = 'сума,'
        self.sheet['P17'] = 'грн.'
        self.sheet['S17'] = 'внесок'
        self.sheet['T17'] = 'внесок'
        self.sheet['U17'] = 'мання'
        self.sheet['W17'] = 'працівни-'
        self.sheet['X17'] = 'періоди'
        self.sheet['Y17'] = 'плати'
        self.sheet['Z17'] = 'виплати'
        self.sheet['AA17'] = 'видачі'
        self.sheet['AB17'] = 'отриманні'
        self.sheet['AC17'] = 'по батькові,'

        self.sheet['F18'] = '(посад.'
        self.sheet['H18'] = 'мат.'
        self.sheet['I18'] = 'мат. доп.'
        self.sheet['O18'] = 'грн.'
        self.sheet['W18'] = 'ком'
        self.sheet['AC18'] = 'ІПН'

        self.sheet['F19'] = 'окладами)'
        self.sheet['H19'] = 'допомога'
        self.sheet['I19'] = '(ВВ)'

        start_column = 1

        for row in self.sheet.iter_rows(min_row=12, max_row=20, min_col=1, max_col=self.sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in range(start_column, self.sheet.max_column):
            cell = self.sheet.cell(row=20, column=column)
            cell.value = column - start_column + 1
            cell.alignment = Alignment(horizontal='center', vertical='center')

        last_row = 21

        for num, worker in enumerate(self.Employer.workers):
            if_feel = self._fill_table(num, worker, last_row)
            if if_feel:
                last_row += 1

        self._fill_last_row(last_row)

        last_row += 3

        self.sheet[f'A{last_row}'] = 'Начальник підрозділу'
        self.sheet[f'F{last_row}'] = f'{self.Employer.name}'
        self.sheet[f'J{last_row}'] = 'Головний бухгалтер'
        self.sheet[f'O{last_row}'] = f'{self.Employer.name}'

        last_row += 1

        self._merge(f'D{last_row}:G{last_row}', f'D{last_row}', '(підпис, прізвище)')
        self._merge(f'K{last_row}:N{last_row}', f'K{last_row}', '(підпис, прізвище)')

        font = Font(name='Arial', bold=False, italic=False, size=9)
        for col in self.sheet.columns:
            for cell in col:
                cell.font = font

        self.sheet['A1'].font = Font(name='Arial', underline='single', size=10)
        #self.sheet['E2'].font = Font(name='Arial', bold=True)
        self.sheet['B10'].font = Font(name='Arial', bold=True, size=11)

        self.sheet.column_dimensions['A'].width = 4.17
        self.sheet.column_dimensions['B'].width = 11.67
        self.sheet.column_dimensions['C'].width = 10.67
        self.sheet.column_dimensions['D'].width = 22.33
        self.sheet.column_dimensions['E'].width = 8.5
        self.sheet.column_dimensions['F'].width = 9.5
        self.sheet.column_dimensions['G'].width = 9.33
        self.sheet.column_dimensions['H'].width = 9.83
        self.sheet.column_dimensions['I'].width = 10.33
        self.sheet.column_dimensions['J'].width = 6.83
        self.sheet.column_dimensions['K'].width = 7.33
        self.sheet.column_dimensions['L'].width = 7.33
        self.sheet.column_dimensions['M'].width = 6.33
        self.sheet.column_dimensions['N'].width = 3.83
        self.sheet.column_dimensions['O'].width = 10.33
        self.sheet.column_dimensions['P'].width = 11.83
        self.sheet.column_dimensions['Q'].width = 9.83
        self.sheet.column_dimensions['R'].width = 8.67
        self.sheet.column_dimensions['S'].width = 8.83
        self.sheet.column_dimensions['T'].width = 8.17
        self.sheet.column_dimensions['U'].width = 16.83
        self.sheet.column_dimensions['V'].width = 8.67
        self.sheet.column_dimensions['W'].width = 9.33
        self.sheet.column_dimensions['X'].width = 10.67
        self.sheet.column_dimensions['Y'].width = 10.67
        self.sheet.column_dimensions['Z'].width = 10.33
        self.sheet.column_dimensions['AA'].width = 10.33
        self.sheet.column_dimensions['AB'].width = 9.33
        self.sheet.column_dimensions['AC'].width = 20.83

        self.sheet.row_dimensions[1].height = 14.25
        self.sheet.row_dimensions[2].height = 15
        self.sheet.row_dimensions[3].height = 16.5
        self.sheet.row_dimensions[4].height = 12
        self.sheet.row_dimensions[5].height = 15
        self.sheet.row_dimensions[6].height = 15
        self.sheet.row_dimensions[7].height = 15
        self.sheet.row_dimensions[8].height = 15
        self.sheet.row_dimensions[9].height = 12
        self.sheet.row_dimensions[10].height = 15
        self.sheet.row_dimensions[11].height = 12
        self.sheet.row_dimensions[12].height = 12
        self.sheet.row_dimensions[13].height = 48
        self.sheet.row_dimensions[14].height = 12
        self.sheet.row_dimensions[15].height = 12
        self.sheet.row_dimensions[16].height = 12
        self.sheet.row_dimensions[17].height = 12
        self.sheet.row_dimensions[18].height = 12
        self.sheet.row_dimensions[19].height = 12
        self.sheet.row_dimensions[20].height = 12
        #self.sheet.row_dimensions[21].height = 36

        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        left_right_border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000')
        )

        left_right_top_border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000')
        )

        bottom_border = Border(
            bottom=Side(border_style='thin', color='000000')
        )

        border_2px = Border(
            left=Side(border_style=BORDER_MEDIUM, color='000000'),
            right=Side(border_style=BORDER_MEDIUM, color='000000'),
            top=Side(border_style=BORDER_MEDIUM, color='000000'),
            bottom=Side(border_style=BORDER_MEDIUM, color='000000')
        )

        self.sheet['E2'].border = border_2px
        self.sheet['F2'].border = border_2px
        self.sheet['C3'].border = bottom_border
        self.sheet['D3'].border = bottom_border
        self.sheet['E3'].border = bottom_border
        self.sheet['F3'].border = bottom_border
        self.sheet['V12'].border = left_right_top_border
        self.sheet['W12'].border = left_right_top_border
        self.sheet['X12'].border = border
        self.sheet['Y12'].border = border
        self.sheet['Z12'].border = border
        self.sheet['AA12'].border = left_right_top_border
        self.sheet['AB12'].border = left_right_top_border
        self.sheet['AC12'].border = left_right_top_border
        self.sheet['M16'].border = left_right_top_border
        self.sheet['N16'].border = left_right_top_border
        self.sheet['O16'].border = left_right_top_border
        self.sheet[f'D{last_row-1}'].border = bottom_border
        self.sheet[f'E{last_row-1}'].border = bottom_border
        self.sheet[f'F{last_row-1}'].border = bottom_border
        self.sheet[f'G{last_row - 1}'].border = bottom_border
        self.sheet[f'J{last_row - 1}'].border = bottom_border
        self.sheet[f'K{last_row-1}'].border = bottom_border
        self.sheet[f'M{last_row-1}'].border = bottom_border
        self.sheet[f'N{last_row-1}'].border = bottom_border
        self.sheet[f'O{last_row-1}'].border = bottom_border

        for row in self.sheet.iter_rows(min_row=13, max_row=19, min_col=11, max_col=11):
            for cell in row:
                cell.border = left_right_border

        for row in self.sheet.iter_rows(min_row=13, max_row=15, min_col=12, max_col=14):
            for cell in row:
                cell.border = left_right_border

        for row in self.sheet.iter_rows(min_row=13, max_row=19, min_col=16, max_col=17):
            for cell in row:
                cell.border = left_right_border

        for row in self.sheet.iter_rows(min_row=13, max_row=19, min_col=18, max_col=29):
            for cell in row:
                cell.border = left_right_border

        for row in self.sheet.iter_rows(min_row=13, max_row=19, min_col=18, max_col=18):
            for cell in row:
                cell.border = border

        for row in self.sheet.iter_rows(min_row=12, max_row=12, min_col=1, max_col=5):
            for cell in row:
                cell.border = left_right_top_border

        for row in self.sheet.iter_rows(min_row=12, max_row=12, min_col=6, max_col=21):
            for cell in row:
                cell.border = border

        for row in self.sheet.iter_rows(min_row=13, max_row=19, min_col=1, max_col=5):
            for cell in row:
                cell.border = left_right_border

        for row in self.sheet.iter_rows(min_row=16, max_row=16, min_col=6, max_col=10):
            for cell in row:
                cell.border = left_right_top_border

        for row in self.sheet.iter_rows(min_row=17, max_row=19, min_col=6, max_col=10):
            for cell in row:
                cell.border = left_right_border

        for row in self.sheet.iter_rows(min_row=17, max_row=19, min_col=12, max_col=14):
            for cell in row:
                cell.border = left_right_border

        for row in self.sheet.iter_rows(min_row=20, max_row=21, min_col=1, max_col=29):
            for cell in row:
                cell.border = border

    def get_bytes(self) -> io.BytesIO:
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def save(self):
        self.workbook.save("example.xlsx")


if __name__ == '__main__':


    worker1 = Worker(
        sex="М",
        name="John Doe",
        job_title="Software Engineer",
        salary="5000",
        working_hours="8",
        ident_IPN="1234567890",
        employment_date="15.09.2023",
        birthday="13.03.1992",
        dismissal=''
    )

    worker2 = Worker(
        sex="Ж",
        name="Jane Smith",
        job_title="Data Scientist",
        salary="6000",
        working_hours="8",
        ident_IPN="0987654321",
        employment_date="18.07.2023",
        birthday="13.03.1992",
        dismissal=''
    )

    # Создание объекта Employer с несколькими работниками
    employer = Employer(
        name="Acme Corporation",
        ident_EDRPOU="123456789",
        workers=[worker1, worker2],
        residence = 'UK',
        phone = '09094594095'
    )

    sheet = SettlementPayment(employer)
    sheet.save()
