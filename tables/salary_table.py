from datetime import timedelta, datetime
from typing import TypedDict

import pandas as pd
from openpyxl.styles import Alignment, Border, Side, Font

from Constants import MONTHS1
from tables.Exceptions import NoSuitableEmployers
from tables.Models import Employer, Worker
from tables.Table import Table


class PartParameters(TypedDict):
    salary: float
    esv: float
    pdfo: float
    military_tax: float
    received: float


class Tax(TypedDict):
    esv: float
    pdfo: float
    military_tax: float


class SalaryTable(Table):
    def __init__(self, employer: Employer):
        super().__init__(employer)
        # self.dates_1_15, self.dates_16_1 = self.get_month_data()

        self.together_parameters = PartParameters(salary=0.0,
                                                  esv=0.0,
                                                  pdfo=0.0,
                                                  military_tax=0.0,
                                                  received=0.0)
        self.tax = Tax(esv=0.0,
                       pdfo=0.0,
                       military_tax=0.0)

        self.working_period = []
        start_day = datetime.now().replace(day=1)
        self.working_period.append(start_day)

        end_day = (start_day + timedelta(days=35)).replace(day=1) - timedelta(days=1)

        self.working_period.append(end_day)

        self.days_per_period = self.count_working_days(self.working_period[0], self.working_period[1])

        self.salary_per_day: float = 0.0

        self.full_salary_workers = []
        self.half_salary_workers = []

        self.__assemble_workbook()

    def get_last_month(self):
        today = datetime.now()
        first_day_of_current_month = today.replace(day=1)
        return first_day_of_current_month - timedelta(days=1)

    def get_month_data(self):
        today = datetime.now()
        first_day_of_current_month = today.replace(day=1)

        # Создание списка с 1 по 15 число нынешнего месяца
        start_date_1 = first_day_of_current_month
        end_date_1 = first_day_of_current_month.replace(day=15)
        list_1 = self.generate_date_list(start_date_1, end_date_1)

        # Создание списка с 16 числа нынешнего месяца по 1 число следующего месяца
        start_date_2 = first_day_of_current_month.replace(day=16)
        end_date_2 = (first_day_of_current_month.replace(day=1) + timedelta(days=13)).replace(day=1)
        list_2 = self.generate_date_list(start_date_2, end_date_2)

        return list_1, list_2

    @staticmethod
    def generate_date_list(start_date, end_date):
        date_list = []
        current_date = start_date
        while current_date <= end_date:
            date_list.append(current_date)
            current_date += timedelta(days=1)
        return date_list

    def count_working_days(self, start_date, end_date):
        # Создаем диапазон дат с использованием библиотеки pandas
        date_range = pd.date_range(start=start_date, end=end_date)

        # Фильтруем только рабочие дни (понедельник - пятница)
        working_days = date_range[date_range.weekday < 5]

        # Возвращаем количество рабочих дней
        return len(working_days)

    def _fill_table(self, num: int, worker: Worker, start_row: int):
        pass

    def _fill_first_part(self, worker: Worker, start_row: int):
        if worker.if_employment_later_last_month():
            return False

        # Получаем количество дней с 1го по 15тое число текущего месяца
        today = datetime.now()
        start_date = today.replace(day=1)
        end_date = start_date.replace(day=15)

        if end_date < worker.get_employment_date:
            end_date = worker.get_employment_date

        if end_date < start_date:
            return False

        payment_date = f'Платимо {today.replace(day=20).strftime("%d.%m.%Y")}'

        name = worker.name

        period = f'Зарплата за першу половину {MONTHS1[today.month-1]}'

        period_dates = f'з {start_date.day} по {end_date.day} {MONTHS1[today.month-1]}'

        count_working_days = self.count_working_days(start_date, end_date)

        salary_per_day = worker.salary_per_day(start_date, end_date)

        salary = worker.salary_for_period(start_date, end_date)
        self.together_parameters['salary'] += salary

        esv = round(salary * 0.22, 2)
        self.together_parameters['esv'] += esv

        pdfo = round(salary * 0.18, 2)
        self.together_parameters['pdfo'] += pdfo

        military_tax = round(salary * 0.015, 2)
        self.together_parameters['military_tax'] += military_tax

        received = round(salary - pdfo - military_tax, 2)
        self.together_parameters['received'] += received

        self.sheet[f'A{start_row}'] = payment_date
        self.sheet[f'B{start_row}'] = name
        self.sheet[f'C{start_row}'] = period
        self.sheet[f'D{start_row}'] = period_dates
        self.sheet[f'E{start_row}'] = f'{count_working_days} роб. днів'
        self.sheet[f'F{start_row}'] = f'{salary_per_day}'
        self.sheet[f'G{start_row}'] = f'{salary}'
        self.sheet[f'H{start_row}'] = f'{esv}'
        self.sheet[f'I{start_row}'] = f'{pdfo}'
        self.sheet[f'J{start_row}'] = f'{military_tax}'
        self.sheet[f'K{start_row}'] = f'{received}'

        return True

    def _fill_second_part(self, worker: Worker, start_row: int):
        if worker.if_employment_later_last_month():
            return False

        # Получаем количество дней с 16го по последнее число текущего месяца
        today = datetime.now()
        start_date = today.replace(day=16)
        end_date = (start_date.replace(day=1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)

        if end_date < worker.get_employment_date:
            end_date = worker.get_employment_date

        if end_date < start_date:
            return False

        # 5тое число следующего месяца
        payment_date = f'Платимо {(today.replace(day=30) + timedelta(days=7)).replace(day=5).strftime("%d.%m.%Y")}'

        name = worker.name

        period = f'Зарплата за другу половину {MONTHS1[today.month - 1]}'

        period_dates = f'з {start_date.day} по {end_date.day} {MONTHS1[today.month - 1]}'

        count_working_days = self.count_working_days(start_date, end_date)

        salary_per_day = worker.salary_per_day(start_date, end_date)

        salary = worker.salary_for_period(start_date, end_date)
        self.together_parameters['salary'] += salary

        esv = round(salary * 0.22, 2)
        self.together_parameters['esv'] += esv

        pdfo = round(salary * 0.18, 2)
        self.together_parameters['pdfo'] += pdfo

        military_tax = round(salary * 0.015, 2)
        self.together_parameters['military_tax'] += military_tax

        received = round(salary - pdfo - military_tax, 2)
        self.together_parameters['received'] += received

        self.sheet[f'A{start_row}'] = payment_date
        self.sheet[f'B{start_row}'] = name
        self.sheet[f'C{start_row}'] = period
        self.sheet[f'D{start_row}'] = period_dates
        self.sheet[f'E{start_row}'] = f'{count_working_days} роб. днів'
        self.sheet[f'F{start_row}'] = f'{salary_per_day}'
        self.sheet[f'G{start_row}'] = f'{salary}'
        self.sheet[f'H{start_row}'] = f'{esv}'
        self.sheet[f'I{start_row}'] = f'{pdfo}'
        self.sheet[f'J{start_row}'] = f'{military_tax}'
        self.sheet[f'K{start_row}'] = f'{received}'

        return True

    def _fil_together_row(self, start_row):
        self.sheet[f'F{start_row}'] = f'Разом:'
        self.sheet[f'G{start_row}'] = f'{self.together_parameters["salary"]}'
        self.sheet[f'H{start_row}'] = f'{self.together_parameters["esv"]}'
        self.sheet[f'I{start_row}'] = f'{self.together_parameters["pdfo"]}'
        self.sheet[f'J{start_row}'] = f'{round(self.together_parameters["military_tax"], 2)}'
        self.sheet[f'K{start_row}'] = f'{round(self.together_parameters["received"], 2)}'

    def _salary_header(self, last_row, worker):
        self.sheet[f'E{last_row}'] = f'{self.days_per_period} роб. днів'

        salary = float(worker.salary_real)
        self.salary_per_day = round(salary/self.days_per_period, 2)
        esv = salary * 0.22
        pdfo = salary * 0.18
        military_tax = salary * 0.015
        self.sheet[f'G{last_row}'] = f'{salary}'
        self.sheet[f'H{last_row}'] = f'{esv}'
        self.sheet[f'I{last_row}'] = f'{pdfo}'
        self.sheet[f'J{last_row}'] = f'{military_tax}'
        self.sheet[f'K{last_row}'] = f'{salary - pdfo - military_tax}'

    def __assemble_workbook(self):
        self.sheet = self.workbook.active

        self.sheet['A1'] = f'ФОП {self.Employer.name}'

        self.sheet['A5'] = 'Дата виплати'
        self.sheet['B5'] = 'ПІБ робітника'
        self._merge('C5:D5', 'C5', 'Виплата за термін')
        self.sheet['E5'] = 'Кількість робочих днів'
        self.sheet['F5'] = 'Оклад за 1 роб. день.'
        self.sheet['G5'] = 'Оклад'
        self.sheet['H5'] = 'ЕСВ'
        self.sheet['I5'] = 'ПДФО'
        self.sheet['J5'] = 'Військовий збір'
        self.sheet['K5'] = 'На руки'

        last_row = 6

        for worker in self.Employer.workers:
            if not worker.if_dismissal_later():
                if worker.working_hours == '8':
                    self.full_salary_workers.append(worker)
                elif worker.working_hours == '4':
                    self.half_salary_workers.append(worker)

        if len(self.full_salary_workers) + len(self.half_salary_workers) == 0:
            raise NoSuitableEmployers(self.Employer)

        if self.full_salary_workers:
            self._salary_header(last_row, self.full_salary_workers[0])

            last_row += 1

            for num, worker in enumerate(self.full_salary_workers):
                if_feel = self._fill_first_part(worker, last_row)
                if if_feel:
                    last_row += 1

            if last_row > 7:
                self._merge(f'F6:F{last_row-1}', 'F6', f'{self.salary_per_day}')
                self._fil_together_row(last_row)

                for cells in self.sheet.iter_cols(min_col=1, max_col=self.sheet.max_column,
                                                  min_row=last_row, max_row=last_row):
                    for cell in cells:
                        cell.font = Font(bold=True)

                last_row += 1

            self.together_parameters = PartParameters(salary=0.0,
                                                      esv=0.0,
                                                      pdfo=0.0,
                                                      military_tax=0.0,
                                                      received=0.0)
            start_second_part_row = last_row

            for num, worker in enumerate(self.full_salary_workers):
                if_feel = self._fill_second_part(worker, last_row)
                if if_feel:
                    last_row += 1

            if start_second_part_row != last_row:
                self._merge(f'F{start_second_part_row}:F{last_row - 1}',
                            f'F{start_second_part_row}',
                            f'{self.salary_per_day}')
                self._fil_together_row(last_row)
                for cells in self.sheet.iter_cols(min_col=1, max_col=self.sheet.max_column,
                                                  min_row=last_row, max_row=last_row):
                    for cell in cells:
                        cell.font = Font(bold=True)

                last_row += 2

        if self.half_salary_workers:
            self._salary_header(last_row, self.half_salary_workers[0])

            last_row += 1

            for num, worker in enumerate(self.half_salary_workers):
                if_feel = self._fill_first_part(worker, last_row)
                if if_feel:
                    last_row += 1

            if last_row > 7:
                self._merge(f'F6:F{last_row - 1}', 'F6', f'{self.salary_per_day}')
                self._fil_together_row(last_row)
                for cells in self.sheet.iter_cols(min_col=1, max_col=self.sheet.max_column,
                                                  min_row=last_row, max_row=last_row):
                    for cell in cells:
                        cell.font = Font(bold=True)

                last_row += 1

            self.together_parameters = PartParameters(salary=0.0,
                                                      esv=0.0,
                                                      pdfo=0.0,
                                                      military_tax=0.0,
                                                      received=0.0)
            start_second_part_row = last_row

            for num, worker in enumerate(self.half_salary_workers):
                if_feel = self._fill_second_part(worker, last_row)
                if if_feel:
                    last_row += 1

            if start_second_part_row != last_row:
                self._merge(f'F{start_second_part_row}:F{last_row - 1}',
                            f'F{start_second_part_row}',
                            f'{self.salary_per_day}')
                self._fil_together_row(last_row)
                for cells in self.sheet.iter_cols(min_col=1, max_col=self.sheet.max_column,
                                                  min_row=last_row, max_row=last_row):
                    for cell in cells:
                        cell.font = Font(bold=True)

                last_row += 2


        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        for cells in self.sheet.iter_cols(min_col=1, max_col=self.sheet.max_column,
                                          min_row=5, max_row=self.sheet.max_row):
            for cell in cells:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

        self.sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        self.sheet.column_dimensions['A'].width = 20
        self.sheet.column_dimensions['B'].width = 23
        self.sheet.column_dimensions['C'].width = 20
        self.sheet.column_dimensions['D'].width = 15
        self.sheet.column_dimensions['E'].width = 15
        self.sheet.column_dimensions['F'].width = 12
        self.sheet.column_dimensions['G'].width = 10
        self.sheet.column_dimensions['I'].width = 10
        self.sheet.column_dimensions['J'].width = 12

        self.sheet.row_dimensions[1].height = 30

        for row_num in range(5, self.sheet.max_row + 1):
            self.sheet.row_dimensions[row_num].height = 35

        self.sheet.row_dimensions[5].height = 55


if __name__ == "__main__":
    worker1 = Worker(
        sex="Male",
        name="John Doe",
        job_title="Software Engineer",
        salary="7100",
        working_hours="8",
        ident_IPN="1234567890",
        employment_date='01.11.2023',
        birthday='01.01.1990',
        dismissal=''
    )

    worker2 = Worker(
        sex="Female",
        name="Jane Smith",
        job_title="Data Scientist",
        salary="7100",
        working_hours="8",
        ident_IPN="0987654321",
        employment_date='01.11.2023',
        birthday='01.01.1990',
        dismissal=''
    )

    # Создание объекта Employer с несколькими работниками
    employer = Employer(
        name="Acme Corporation",
        ident_EDRPOU="123456789",
        workers=[worker1, worker2],
        residence='qwee',
        phone='12312312'
    )

    table = SalaryTable(employer)
    table.save('testSalaryTable')
