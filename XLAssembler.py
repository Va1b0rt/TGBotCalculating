import datetime
import io

import calendar
import itertools
import re
from typing import Union

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Side
from openpyxl.styles import Font, Border, PatternFill, Alignment

from logger import Logger

cls_logger = Logger()
logger = cls_logger.get_logger

months_names: list[str] = ['січень', 'лютий', 'березень',
                           'квітень', 'травень', 'червень',
                           'липень', 'серпень', 'вересень',
                           'жовтень', 'листопад', 'грудень']


@logger.catch
def get_quarter(month: int) -> str:
    if month <= 3:
        return '1'
    elif month in (4, 5, 6):
        return '2'
    elif month in (7, 8, 9):
        return '3'
    else:
        return '4'


@logger.catch
def set_day_in_date(new_day: str, new_month: str, _year: str) -> str:
    year = int(_year)
    month = int(new_month)
    new_date = datetime.datetime(year, month, int(new_day))
    return new_date.strftime('%d.%m.%Y')


@logger.catch
def get_days_of_month(month: int, year: int) -> list[str]:
    year = year
    _, num_days = calendar.monthrange(year, month)
    days = [str(day).zfill(2) for day in range(1, num_days + 1)]
    return days


class TableAssembler:
    def __init__(self, raw_body: dict[str, Union[float, str, list[str]]]):
        self.months_count: list[str] = []
        self.years_actual: list[str] = []
        self.raw_body = raw_body
        self.tittle = raw_body['tittle']
        self.workbooks: list[Workbook] = []
        self._set_month_count()
        self._set_years_count()
        self.fop_sums: list[list[str]] = []
        self.all_months_sum = 0

        self.__assemble_workbook()

    @logger.catch
    def _set_month_count(self):

        for key in list(self.raw_body.keys()):
            if re.search(r'^\d*.\d*.\d*$', key):

                if self.months_count and (key.split('.')[1] if len(key) > 2 else key) in self.months_count:
                    continue

                self.months_count.append(key.split('.')[1] if len(key) > 2 else key)

        self.months_count = sorted(self.months_count, key=lambda x: int(x))


    @logger.catch
    def _set_years_count(self):
        for key in list(self.raw_body.keys()):
            if re.search(r'^\d*\.\d*\.\d*$', key) and len(key) == 10:
                if self.years_actual and key.split('.')[2] in self.years_actual:
                    continue

                self.years_actual.append(key.split('.')[2])

    @logger.catch
    def __set_fop_sums(self, fop_sums: list[str]):
        words = self.tittle.lower().replace('фоп', '').split(' ')
        all_combinations = list(itertools.permutations(words))

        for combinate in all_combinations:
            for num, fop in enumerate(fop_sums):
                try:
                    if f'{combinate[0]} {combinate[1]} {combinate[2]}' in fop.lower() or f'{combinate[0]} {combinate[1][:1]} {combinate[2][:1]}' in fop.lower() or f'{combinate[0]} {combinate[1][:1]}. {combinate[2][:1]}.' in fop.lower():
                        del fop_sums[num]
                except IndexError as er:
                    #print(er)
                    pass

        self.fop_sums.append(fop_sums)

    @logger.catch
    def _month_in_this_year(self, control_month: str, control_year: str):
        pattern = re.compile(r'\d*\.\d*\.\d*')
        for key in list(self.raw_body.keys()):
            if pattern.search(key):
                if control_month in key.split('.')[1]:
                    if control_year in key.split('.')[2]:
                        return True
        return False

    @logger.catch
    def __assemble_workbook(self):
        self.workbook_count = 0
        for actual_year in self.years_actual:
            for num, month in enumerate(self.months_count):
                if not self._month_in_this_year(month, actual_year):
                    continue

                self.workbooks.append(Workbook())
                sheet: Worksheet = self.workbooks[self.workbook_count].active

                # ROWS/COLUMNS OPTIONS

                sheet.column_dimensions['A'].width = 3
                sheet.column_dimensions['B'].width = 10
                sheet.column_dimensions['C'].width = 4
                sheet.column_dimensions['D'].width = 10
                sheet.column_dimensions['E'].width = 12
                sheet.column_dimensions['F'].width = 15
                sheet.column_dimensions['G'].width = 15
                sheet.column_dimensions['H'].width = 15
                sheet.column_dimensions['I'].width = 15
                sheet.column_dimensions['J'].width = 15
                sheet.column_dimensions['K'].width = 10
                sheet.column_dimensions['L'].width = 10

                sheet.row_dimensions[9].height = 35
                sheet.row_dimensions[10].height = 25
                sheet.row_dimensions[11].height = 70

                # HEADER
                sheet['B1'] = self.tittle

                font = Font(name='Arial', bold=True, italic=False, size=12)
                sheet['B5'] = 'Книга обліку доходів '

                sheet['B5'].font = font
                sheet.merge_cells('B5:L5')
                alignment_center = Alignment(horizontal='center', vertical='center', wrapText=True)
                sheet['B5'].alignment = alignment_center

                sheet.merge_cells('B6:L6')
                sheet[
                    'B6'] = '(для платників єдиного податку першої і другої груп та платників єдиного податку третьої групи,'
                sheet['B6'].alignment = alignment_center
                sheet['B6'].font = Font(name='Arial', bold=False, italic=False, underline='single', size=12)

                sheet.merge_cells('B7:L7')
                sheet['B7'] = 'які не є платниками податку на додану вартість)'
                sheet['B7'].alignment = alignment_center
                sheet['B7'].font = Font(name='Arial', bold=False, italic=False, underline='single', size=12)

                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )

                table_font = Font(name='Arial', bold=False, italic=False, size=10)

                sheet['B9'].border = border
                sheet['B9'].font = table_font
                sheet['B9'].alignment = alignment_center
                sheet['B9'] = 'Дата'
                sheet.merge_cells('B9:D11')

                sheet['E9'].border = border
                sheet['E9'].font = table_font
                sheet['E9'].alignment = alignment_center
                sheet['E9'] = 'Дохід від провадження діяльності*'
                sheet.merge_cells('E9:J9')

                sheet['K9'].border = border
                sheet['K9'].font = table_font
                sheet['K9'].alignment = alignment_center
                sheet['K9'] = 'Дохід, що оподатковується за ставкою 15% **'
                sheet.merge_cells('K9:L9')

                sheet['E10'].border = border
                sheet['E10'].font = table_font
                sheet['E10'].alignment = alignment_center
                sheet['E10'] = 'вартість проданих товарів, виконаних робіт, наданих послуг'
                sheet.merge_cells('E10:G10')

                sheet['E11'].border = border
                sheet['E11'].font = table_font
                sheet['E11'].alignment = alignment_center
                sheet['E11'] = 'сума, грн.'

                sheet['F11'].border = border
                sheet['F11'].font = table_font
                sheet['F11'].alignment = alignment_center
                sheet['F11'] = 'сума повернутих коштів за товар (роботи, послуги) та/або передплати, грн.'

                sheet['G11'].border = border
                sheet['G11'].font = table_font
                sheet['G11'].alignment = alignment_center
                sheet['G11'] = 'скоригована сума доходу, грн. (гр.2 – гр.3)'

                sheet['H10'].border = border
                sheet['H10'].font = table_font
                sheet['H10'].alignment = alignment_center
                sheet['H10'] = 'вартість безоплатно отриманих товарів (робіт, послуг), грн.'
                sheet.merge_cells('H10:H11')

                sheet['I10'].border = border
                sheet['I10'].font = table_font
                sheet['I10'].alignment = alignment_center
                sheet['I10'] = 'сума заборгованості, за якою минув строк позовної давності, грн.'
                sheet.merge_cells('I10:I11')

                sheet['J10'].border = border
                sheet['J10'].font = table_font
                sheet['J10'].alignment = alignment_center
                sheet['J10'] = 'всього, грн. (гр.4 + гр.5 + гр.6)'
                sheet.merge_cells('J10:J11')

                sheet['K10'].border = border
                sheet['K10'].font = table_font
                sheet['K10'].alignment = alignment_center
                sheet['K10'] = 'вид доходу'
                sheet.merge_cells('K10:K11')

                sheet['L10'].border = border
                sheet['L10'].font = table_font
                sheet['L10'].alignment = alignment_center
                sheet['L10'] = 'сума, грн.'
                sheet.merge_cells('L10:L11')

                sheet['B12'].border = border
                sheet['B12'].font = table_font
                sheet['B12'].alignment = alignment_center
                sheet['B12'] = '1'
                sheet.merge_cells('B12:D12')

                sheet['E12'].border = border
                sheet['E12'].font = table_font
                sheet['E12'].alignment = alignment_center
                sheet['E12'] = '2'

                sheet['F12'].border = border
                sheet['F12'].font = table_font
                sheet['F12'].alignment = alignment_center
                sheet['F12'] = '3'

                sheet['G12'].border = border
                sheet['G12'].font = table_font
                sheet['G12'].alignment = alignment_center
                sheet['G12'] = '4'

                sheet['H12'].border = border
                sheet['H12'].font = table_font
                sheet['H12'].alignment = alignment_center
                sheet['H12'] = '5'

                sheet['I12'].border = border
                sheet['I12'].font = table_font
                sheet['I12'].alignment = alignment_center
                sheet['I12'] = '6'

                sheet['J12'].border = border
                sheet['J12'].font = table_font
                sheet['J12'].alignment = alignment_center
                sheet['J12'] = '7'

                sheet['K12'].border = border
                sheet['K12'].font = table_font
                sheet['K12'].alignment = alignment_center
                sheet['K12'] = '8'

                sheet['L12'].border = border
                sheet['L12'].font = table_font
                sheet['L12'].alignment = alignment_center
                sheet['L12'] = '9'

                left_border = Border(
                    left=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )

                top_bot_border = Border(
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )

                right_border = Border(
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )

                # BODY

                all_in_font = Font(name='Arial', bold=False, italic=True, size=10)

                # Строка на которой закончилась итерация
                last_row: int = 13

                if not list(self.raw_body.keys()):
                    return

                def get_key(keys_lst: list[str]) -> str:
                    pattern = re.compile(r'\d*.\d*.\d*')
                    for key in keys_lst:
                        if pattern.search(key):
                            return key




                actual_date = datetime.datetime.strptime(get_key(list(self.raw_body.keys())), '%d.%m.%Y')

                days: list[str] = get_days_of_month(int(month), int(actual_year))

                row_num: list[int] = list(range(13, len(days) + 14))

                for day, row_num in zip(days, row_num):
                    day_now = set_day_in_date(day, month, actual_year)

                    try:
                        if day_now.split('.')[1] != month:
                            continue
                    except:
                        pass

                    if_contains = True if day_now in self.raw_body else False

                    sheet[f'B{str(row_num)}'].border = left_border
                    sheet[f'B{str(row_num)}'].font = all_in_font
                    sheet[f'B{str(row_num)}'].alignment = alignment_center
                    sheet[f'B{str(row_num)}'] = 'Разом за'

                    sheet[f'C{str(row_num)}'].border = top_bot_border
                    sheet[f'C{str(row_num)}'].font = table_font
                    sheet[f'C{str(row_num)}'].alignment = alignment_center
                    sheet[f'C{str(row_num)}'] = day

                    sheet[f'D{str(row_num)}'].border = right_border
                    sheet[f'D{str(row_num)}'].font = table_font
                    sheet[f'D{str(row_num)}'].alignment = alignment_center
                    sheet[f'D{str(row_num)}'] = ''  # day_now if if_contains else ''

                    sheet[f'E{str(row_num)}'].border = border
                    sheet[f'E{str(row_num)}'].font = table_font
                    sheet[f'E{str(row_num)}'].alignment = alignment_center
                    sheet[f'E{str(row_num)}'] = self.raw_body[day_now] if if_contains else '0,00'

                    sheet[f'F{str(row_num)}'].border = border
                    sheet[f'F{str(row_num)}'].font = table_font
                    sheet[f'F{str(row_num)}'].alignment = alignment_center
                    sheet[f'F{str(row_num)}'] = ''

                    sheet[f'G{str(row_num)}'].border = border
                    sheet[f'G{str(row_num)}'].font = table_font
                    sheet[f'G{str(row_num)}'].alignment = alignment_center
                    sheet[f'G{str(row_num)}'] = self.raw_body[day_now] if if_contains else '0.00'

                    sheet[f'H{str(row_num)}'].border = border
                    sheet[f'H{str(row_num)}'].font = table_font
                    sheet[f'H{str(row_num)}'].alignment = alignment_center
                    sheet[f'H{str(row_num)}'] = ''

                    sheet[f'I{str(row_num)}'].border = border
                    sheet[f'I{str(row_num)}'].font = table_font
                    sheet[f'I{str(row_num)}'].alignment = alignment_center
                    sheet[f'I{str(row_num)}'] = ''

                    sheet[f'J{str(row_num)}'].border = border
                    sheet[f'J{str(row_num)}'].font = table_font
                    sheet[f'J{str(row_num)}'].alignment = alignment_center
                    sheet[f'J{str(row_num)}'] = self.raw_body[day_now] if if_contains else '0.00'

                    sheet[f'K{str(row_num)}'].border = border
                    sheet[f'K{str(row_num)}'].font = table_font
                    sheet[f'K{str(row_num)}'].alignment = alignment_center
                    sheet[f'K{str(row_num)}'] = ''

                    sheet[f'L{str(row_num)}'].border = border
                    sheet[f'L{str(row_num)}'].font = table_font
                    sheet[f'L{str(row_num)}'].alignment = alignment_center
                    sheet[f'L{str(row_num)}'] = ''

                    last_row += 1

                # FOOTER
                footer_font = Font(name='Arial', bold=True, italic=True, size=10)


                sheet.row_dimensions[last_row].height = 70

                sheet[f'B{str(last_row)}'].border = border
                sheet[f'B{str(last_row)}'].font = footer_font
                sheet[f'B{str(last_row)}'].alignment = alignment_center
                sheet[f'B{str(last_row)}'] = f'Наростаючим підсумком за {months_names[int(month) - 1]} {actual_year} року:'

                sheet[f'E{str(last_row)}'].border = border
                sheet[f'E{str(last_row)}'].font = footer_font
                sheet[f'E{str(last_row)}'].alignment = alignment_center
                sheet[f'E{str(last_row)}'] = self.raw_body['months'][int(month) - 1]

                sheet[f'F{str(last_row)}'].border = border
                sheet[f'F{str(last_row)}'].font = footer_font
                sheet[f'F{str(last_row)}'].alignment = alignment_center
                sheet[f'F{str(last_row)}'] = ''

                sheet[f'G{str(last_row)}'].border = border
                sheet[f'G{str(last_row)}'].font = footer_font
                sheet[f'G{str(last_row)}'].alignment = alignment_center
                sheet[f'G{str(last_row)}'] = self.raw_body['months'][int(month) - 1]

                sheet[f'H{str(last_row)}'].border = border
                sheet[f'H{str(last_row)}'].font = footer_font
                sheet[f'H{str(last_row)}'].alignment = alignment_center
                sheet[f'H{str(last_row)}'] = ''

                sheet[f'I{str(last_row)}'].border = border
                sheet[f'I{str(last_row)}'].font = footer_font
                sheet[f'I{str(last_row)}'].alignment = alignment_center
                sheet[f'I{str(last_row)}'] = ''

                sheet[f'J{str(last_row)}'].border = border
                sheet[f'J{str(last_row)}'].font = footer_font
                sheet[f'J{str(last_row)}'].alignment = alignment_center
                sheet[f'J{str(last_row)}'] = self.raw_body['months'][int(month) - 1]

                sheet[f'K{str(last_row)}'].border = border
                sheet[f'K{str(last_row)}'].font = footer_font
                sheet[f'K{str(last_row)}'].alignment = alignment_center
                sheet[f'K{str(last_row)}'] = ''

                sheet[f'L{str(last_row)}'].border = border
                sheet[f'L{str(last_row)}'].font = footer_font
                sheet[f'L{str(last_row)}'].alignment = alignment_center
                sheet[f'L{str(last_row)}'] = ''
                sheet.merge_cells(f'B{str(last_row)}:D{str(last_row)}')

                last_row += 1

                blue_fill = PatternFill(fill_type='solid', fgColor='9999FF')


                sheet.row_dimensions[last_row].height = 70

                sheet[f'B{str(last_row)}'].fill = blue_fill
                sheet[f'B{str(last_row)}'].border = border
                sheet[f'B{str(last_row)}'].font = footer_font
                sheet[f'B{str(last_row)}'].alignment = alignment_center
                sheet[f'B{str(last_row)}'] = f'Наростаючим підсумком за {get_quarter(int(month))} квартал {actual_year} року:'

                sheet[f'E{str(last_row)}'].fill = blue_fill
                sheet[f'E{str(last_row)}'].border = border
                sheet[f'E{str(last_row)}'].font = footer_font
                sheet[f'E{str(last_row)}'].alignment = alignment_center
                sheet[f'E{str(last_row)}'] = self.raw_body['months'][int(month) - 1]

                sheet[f'F{str(last_row)}'].fill = blue_fill
                sheet[f'F{str(last_row)}'].border = border
                sheet[f'F{str(last_row)}'].font = footer_font
                sheet[f'F{str(last_row)}'].alignment = alignment_center
                sheet[f'F{str(last_row)}'] = ''

                sheet[f'G{str(last_row)}'].fill = blue_fill
                sheet[f'G{str(last_row)}'].border = border
                sheet[f'G{str(last_row)}'].font = footer_font
                sheet[f'G{str(last_row)}'].alignment = alignment_center
                sheet[f'G{str(last_row)}'] = self.raw_body['months'][int(month) - 1]

                sheet[f'H{str(last_row)}'].fill = blue_fill
                sheet[f'H{str(last_row)}'].border = border
                sheet[f'H{str(last_row)}'].font = footer_font
                sheet[f'H{str(last_row)}'].alignment = alignment_center
                sheet[f'H{str(last_row)}'] = ''

                sheet[f'I{str(last_row)}'].fill = blue_fill
                sheet[f'I{str(last_row)}'].border = border
                sheet[f'I{str(last_row)}'].font = footer_font
                sheet[f'I{str(last_row)}'].alignment = alignment_center
                sheet[f'I{str(last_row)}'] = ''

                sheet[f'J{str(last_row)}'].fill = blue_fill
                sheet[f'J{str(last_row)}'].border = border
                sheet[f'J{str(last_row)}'].font = footer_font
                sheet[f'J{str(last_row)}'].alignment = alignment_center
                sheet[f'J{str(last_row)}'] = self.raw_body['months'][int(month) - 1]

                sheet[f'K{str(last_row)}'].fill = blue_fill
                sheet[f'K{str(last_row)}'].border = border
                sheet[f'K{str(last_row)}'].font = footer_font
                sheet[f'K{str(last_row)}'].alignment = alignment_center
                sheet[f'K{str(last_row)}'] = ''

                sheet[f'L{str(last_row)}'].fill = blue_fill
                sheet[f'L{str(last_row)}'].border = border
                sheet[f'L{str(last_row)}'].font = footer_font
                sheet[f'L{str(last_row)}'].alignment = alignment_center
                sheet[f'L{str(last_row)}'] = ''
                sheet.merge_cells(f'B{str(last_row)}:D{str(last_row)}')

                last_row += 1

                beige_fill = PatternFill(fill_type='solid', fgColor='FFFFCC')

                sheet.row_dimensions[last_row].height = 70

                sheet.merge_cells(f'B{str(last_row)}:D{str(last_row)}')
                sheet[f'B{str(last_row)}'].fill = beige_fill
                sheet[f'B{str(last_row)}'].border = border
                sheet[f'B{str(last_row)}'].font = footer_font
                sheet[f'B{str(last_row)}'].alignment = alignment_center
                sheet[f'B{str(last_row)}'] = f'Наростаючим підсумком за {actual_year} рік:'

                self.all_months_sum += float(self.raw_body['months'][int(month) - 1])

                sheet[f'E{str(last_row)}'].fill = beige_fill
                sheet[f'E{str(last_row)}'].border = border
                sheet[f'E{str(last_row)}'].font = footer_font
                sheet[f'E{str(last_row)}'].alignment = alignment_center
                sheet[f'E{str(last_row)}'] = self.all_months_sum

                sheet[f'F{str(last_row)}'].fill = beige_fill
                sheet[f'F{str(last_row)}'].border = border
                sheet[f'F{str(last_row)}'].font = footer_font
                sheet[f'F{str(last_row)}'].alignment = alignment_center
                sheet[f'F{str(last_row)}'] = ''

                sheet[f'G{str(last_row)}'].fill = beige_fill
                sheet[f'G{str(last_row)}'].border = border
                sheet[f'G{str(last_row)}'].font = footer_font
                sheet[f'G{str(last_row)}'].alignment = alignment_center
                sheet[f'G{str(last_row)}'] = self.all_months_sum

                sheet[f'H{str(last_row)}'].fill = beige_fill
                sheet[f'H{str(last_row)}'].border = border
                sheet[f'H{str(last_row)}'].font = footer_font
                sheet[f'H{str(last_row)}'].alignment = alignment_center
                sheet[f'H{str(last_row)}'] = ''

                sheet[f'I{str(last_row)}'].fill = beige_fill
                sheet[f'I{str(last_row)}'].border = border
                sheet[f'I{str(last_row)}'].font = footer_font
                sheet[f'I{str(last_row)}'].alignment = alignment_center
                sheet[f'I{str(last_row)}'] = ''

                sheet[f'J{str(last_row)}'].fill = beige_fill
                sheet[f'J{str(last_row)}'].border = border
                sheet[f'J{str(last_row)}'].font = footer_font
                sheet[f'J{str(last_row)}'].alignment = alignment_center
                sheet[f'J{str(last_row)}'] = self.all_months_sum

                sheet[f'K{str(last_row)}'].fill = beige_fill
                sheet[f'K{str(last_row)}'].border = border
                sheet[f'K{str(last_row)}'].font = footer_font
                sheet[f'K{str(last_row)}'].alignment = alignment_center
                sheet[f'K{str(last_row)}'] = ''

                sheet[f'L{str(last_row)}'].fill = beige_fill
                sheet[f'L{str(last_row)}'].border = border
                sheet[f'L{str(last_row)}'].font = footer_font
                sheet[f'L{str(last_row)}'].alignment = alignment_center
                sheet[f'L{str(last_row)}'] = ''
                sheet.merge_cells(f'B{str(last_row)}:D{str(last_row)}')

                last_row += 2

                top_border = Border(
                    top=Side(border_style='thin', color='000000')
                )

                font_bottom_footnotes = Font(name='Arial', bold=False, italic=False, size=8)

                sheet[f'B{str(last_row)}'].border = top_border
                sheet[f'B{str(last_row)}'].font = font_bottom_footnotes
                sheet[
                    f'B{str(last_row)}'] = '* Відповідно до пункту 293.2 та підпункту 2 пункту 293.3 статті 293 глави 1 розділу ХIV Податкового кодексу України (крім доходу, що оподатковується за ставкою 15%).'

                sheet[f'C{str(last_row)}'].border = top_border

                sheet[f'D{str(last_row)}'].border = top_border

                sheet[f'E{str(last_row)}'].border = top_border

                last_row += 1

                sheet[f'B{str(last_row)}'].font = font_bottom_footnotes
                sheet[
                    f'B{str(last_row)}'] = '** Відповідно до підпунктів 2–4 пункту 293.4 статті 293 глави 1 розділу ХIV Податкового кодексу України.'

                sheet['M9'].border = Border(left=Side(border_style='thin', color='000000'))

                if month in self.raw_body:
                    self.__set_fop_sums([f'\n{months_names[int(month) - 1]} {actual_year}\n', *self.raw_body[month]])

                self.workbook_count += 1

    @logger.catch
    def save(self, filename):
        for num, wb in enumerate(self.workbooks):
            wb.save(f'{self.months_count[num]}_{filename}')
            print(f"Table saved to {self.months_count[num]}_{filename}")

    @logger.catch
    def get_bytes(self) -> tuple[list[io.BytesIO], Union[io.BytesIO, None]]:
        workbooks_bytes: list[io.BytesIO] = []
        for wb in self.workbooks:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            workbooks_bytes.append(output)

        if len(self.fop_sums) > 0:

            fop_info_bytes: io.BytesIO = io.BytesIO()

            for fop_info in self.fop_sums:
                for line in fop_info:
                    fop_info_bytes.write(f'{line}\n'.encode('utf-8'))
            fop_info_bytes.seek(0)

        else:
            fop_info_bytes = None


        return workbooks_bytes, fop_info_bytes
